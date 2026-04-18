"""
OFX and CSV import service.

Supports: Itaú OFX, Nubank CSV, Inter CSV, BB OFX/CSV, and the custom
reconciled XLSX format (lancamento_por_lancamento_conferido.xlsx).

Design choice: Use ofxparse for OFX (stable, well-maintained).
CSV parsing is hand-rolled for bank-specific column layouts.
XLSX parsing uses openpyxl.
"""
import csv
import io
import re
import uuid
from datetime import date, datetime
from typing import Any

import ofxparse
import openpyxl


_BANK_DETECT: dict[str, list[str]] = {
    "itau": ["itau", "itaú", "banco itau"],
    "nubank": ["nubank", "nu pagamentos"],
    "inter": ["inter", "banco inter"],
    "bb": ["banco do brasil", "bb ", "bradescobanking"],
}


def _detect_bank(header_text: str) -> str:
    """Heuristic bank detection from OFX FI or CSV headers."""
    lower = header_text.lower()
    for bank, hints in _BANK_DETECT.items():
        if any(h in lower for h in hints):
            return bank
    return "unknown"


def _parse_date(s: str) -> date:
    """Parse YYYYMMDD or DD/MM/YYYY date strings."""
    s = s.strip()
    if re.match(r"^\d{8}$", s):
        return datetime.strptime(s, "%Y%m%d").date()
    if re.match(r"^\d{2}/\d{2}/\d{4}$", s):
        return datetime.strptime(s, "%d/%m/%Y").date()
    if re.match(r"^\d{4}-\d{2}-\d{2}", s):
        return datetime.strptime(s[:10], "%Y-%m-%d").date()
    raise ValueError(f"Unknown date format: {s!r}")


def _brl_cents(s: str) -> int:
    """Parse Brazilian currency string to cents. E.g. '-1.234,56' → -123456."""
    s = s.strip().replace("R$", "").replace(" ", "")
    negative = s.startswith("-")
    s = s.lstrip("-+")
    # Remove thousands separator (dot) and replace decimal comma with dot
    s = s.replace(".", "").replace(",", ".")
    cents = round(float(s) * 100)
    return -cents if negative else cents


def parse_ofx(content: bytes) -> list[dict[str, Any]]:
    """Parse OFX file (Itaú or BB format) into transaction dicts."""
    ofx = ofxparse.OfxParser.parse(io.BytesIO(content))
    rows: list[dict[str, Any]] = []

    for account in ofx.accounts:
        for txn in account.statement.transactions:
            amount_raw: float = float(txn.amount)
            amount_cents = round(abs(amount_raw) * 100)
            tx_type = "income" if amount_raw > 0 else "expense"
            rows.append({
                "id": str(uuid.uuid4()),
                "amount_cents": amount_cents,
                "type": tx_type,
                "description": (txn.memo or txn.payee or "").strip()[:200],
                "date": txn.date.date() if hasattr(txn.date, "date") else txn.date,
                "notes": f"OFX ID: {txn.id}",
                "is_reconciled": True,
                "account_id": None,   # caller must assign
                "category_id": None,
            })

    return rows


def parse_nubank_csv(content: bytes) -> list[dict[str, Any]]:
    """
    Parse Nubank credit card CSV export.
    Columns: date,category,title,amount
    """
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows: list[dict[str, Any]] = []

    for row in reader:
        date_str = row.get("date") or row.get("Data") or ""
        title = row.get("title") or row.get("Descrição") or ""
        amount_str = row.get("amount") or row.get("Valor") or "0"

        amount_cents = _brl_cents(amount_str)
        tx_type = "income" if amount_cents > 0 else "expense"

        rows.append({
            "id": str(uuid.uuid4()),
            "amount_cents": abs(amount_cents),
            "type": tx_type,
            "description": title.strip()[:200],
            "date": _parse_date(date_str),
            "notes": None,
            "is_reconciled": True,
            "account_id": None,
            "category_id": None,
        })

    return rows


def parse_inter_csv(content: bytes) -> list[dict[str, Any]]:
    """
    Parse Banco Inter account statement CSV.
    Columns: Data Lançamento,Histórico,Descrição,Valor,Saldo
    """
    text = content.decode("utf-8-sig")
    # Inter uses semicolons
    reader = csv.DictReader(io.StringIO(text), delimiter=";")
    rows: list[dict[str, Any]] = []

    for row in reader:
        date_str = row.get("Data Lançamento") or row.get("Data") or ""
        descricao = (row.get("Descrição") or row.get("Historico") or "").strip()
        amount_str = row.get("Valor") or "0"

        try:
            amount_cents = _brl_cents(amount_str)
        except ValueError:
            continue

        tx_type = "income" if amount_cents > 0 else "expense"

        rows.append({
            "id": str(uuid.uuid4()),
            "amount_cents": abs(amount_cents),
            "type": tx_type,
            "description": descricao[:200],
            "date": _parse_date(date_str),
            "notes": None,
            "is_reconciled": True,
            "account_id": None,
            "category_id": None,
        })

    return rows


def parse_bb_csv(content: bytes) -> list[dict[str, Any]]:
    """
    Parse Banco do Brasil account statement CSV.
    Columns: Data,Histórico,Nro.Docto.,Valor,Saldo do Período
    """
    text = content.decode("latin-1")
    reader = csv.DictReader(io.StringIO(text), delimiter=";")
    rows: list[dict[str, Any]] = []

    for row in reader:
        date_str = row.get("Data") or ""
        hist = (row.get("Histórico") or row.get("Historico") or "").strip()
        amount_str = row.get("Valor") or "0"

        try:
            amount_cents = _brl_cents(amount_str)
        except ValueError:
            continue

        if not date_str.strip():
            continue

        tx_type = "income" if amount_cents > 0 else "expense"

        rows.append({
            "id": str(uuid.uuid4()),
            "amount_cents": abs(amount_cents),
            "type": tx_type,
            "description": hist[:200],
            "date": _parse_date(date_str),
            "notes": None,
            "is_reconciled": True,
            "account_id": None,
            "category_id": None,
        })

    return rows


# ---------------------------------------------------------------------------
# XLSX reconciled import
# ---------------------------------------------------------------------------

# Sheets that contain actual transactions (in the reconciled xlsx format)
_XLSX_BANK_SHEETS: dict[str, str] = {
    "Itaú": "itau",
    "Inter": "inter",
}

# Row tipos that are NOT regular transactions
_XLSX_SKIP_TIPOS = {"Saldo inicial", "Saldo final conferido"}


def _xlsx_cell_to_float(val: Any) -> float:
    """Convert an openpyxl cell value (float, int, or str) to a Python float."""
    if isinstance(val, (int, float)):
        return float(val)
    if isinstance(val, str):
        # BRL string like "-1.234,56" or "1234.56"
        s = val.strip().replace("R$", "").replace(" ", "")
        negative = s.startswith("-")
        s = s.lstrip("-+")
        s = s.replace(".", "").replace(",", ".")
        result = float(s)
        return -result if negative else result
    raise ValueError(f"Cannot convert xlsx cell value {val!r} to float")


def _xlsx_cell_to_date(val: Any) -> date:
    """Convert an openpyxl cell value (datetime, date, or str) to a date."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        return _parse_date(val)
    raise ValueError(f"Cannot convert xlsx cell value {val!r} to date")


def _normalize_bank_key(name: str) -> str:
    """Return a normalised bank key ('itau' or 'inter') from a free-form string."""
    lower = name.lower()
    if "itau" in lower or "itaú" in lower:
        return "itau"
    if "inter" in lower:
        return "inter"
    return lower.strip()


def _is_transfer_source(transfer_field: str, current_bank_key: str) -> bool:
    """Return True if the current bank is the SOURCE (sender) of this transfer."""
    if not transfer_field or "->" not in transfer_field:
        return False
    src = transfer_field.split("->", maxsplit=1)[0].strip()
    return _normalize_bank_key(src) == current_bank_key


def _resolve_transfer_dest(
    transfer_field: str,
    itau_account_id: str,
    inter_account_id: str,
) -> str | None:
    """
    Parse 'Transferência' column (e.g. 'Itaú -> Inter') and return the
    account_id of the DESTINATION account.

    Returns None when the field is empty or unrecognised.
    """
    if not transfer_field or "->" not in transfer_field:
        return None

    dst_key = _normalize_bank_key(transfer_field.split("->", maxsplit=1)[1])

    if dst_key == "itau":
        return itau_account_id
    if dst_key == "inter":
        return inter_account_id
    return None


def _parse_xlsx_sheet(
    ws: Any,  # openpyxl Worksheet
    account_id: str,
    bank_key: str,
    itau_account_id: str,
    inter_account_id: str,
) -> dict[str, Any]:
    """
    Parse one bank sheet from the reconciled xlsx.

    Expected columns (row 1 = header):
        Ordem | Data | Tipo | Categoria | Transferência | Descrição | Valor | Saldo após lançamento

    Returns:
        {
            "rows": list[dict],         — transaction dicts (no 'id', caller assigns)
            "initial_balance_cents": int,
            "expected_final_cents": int, — from 'Saldo final conferido' row
            "calculated_final_cents": int,
            "validation_errors": list[str],
        }
    """
    rows_iter = ws.iter_rows(values_only=True)

    # — skip header row —
    header = next(rows_iter, None)
    if header is None:
        raise ValueError(f"Sheet '{ws.title}' is empty")

    # Locate column indexes by header name (case-insensitive, strip whitespace)
    col_map: dict[str, int] = {}
    for i, h in enumerate(header):
        if h is not None:
            col_map[str(h).strip().lower()] = i

    def _get(row: tuple, name: str) -> Any:
        idx = col_map.get(name)
        return row[idx] if idx is not None and idx < len(row) else None

    # -----------------------------------------------------------------------
    initial_balance_cents: int = 0
    expected_final_cents: int = 0
    calculated_balance: float = 0.0
    transactions: list[dict[str, Any]] = []
    validation_errors: list[str] = []

    all_data_rows = list(rows_iter)

    for row_num, row in enumerate(all_data_rows, start=2):  # row 1 = header
        tipo_raw = _get(row, "tipo")
        if tipo_raw is None:
            continue  # blank row

        tipo = str(tipo_raw).strip()

        # --- SALDO INICIAL (first special row) ------------------------------
        if tipo == "Saldo inicial":
            val = _get(row, "valor")
            saldo_col = _get(row, "saldo após lançamento")
            raw = saldo_col if saldo_col is not None else val
            if raw is not None:
                # Initial balance sets the starting point
                initial_balance_cents = round(_xlsx_cell_to_float(raw) * 100)
                calculated_balance = _xlsx_cell_to_float(raw)
            continue

        # --- SALDO FINAL CONFERIDO (last special row) -----------------------
        if tipo == "Saldo final conferido":
            val = _get(row, "valor")
            saldo_col = _get(row, "saldo após lançamento")
            raw = saldo_col if saldo_col is not None else val
            if raw is not None:
                expected_final_cents = round(_xlsx_cell_to_float(raw) * 100)
            continue

        # --- REGULAR TRANSACTION ROWS ---------------------------------------
        date_val = _get(row, "data")
        valor_val = _get(row, "valor")
        descricao_val = _get(row, "descrição")
        categoria_val = _get(row, "categoria")
        transferencia_val = _get(row, "transferência")
        ordem_val = _get(row, "ordem")

        if date_val is None or valor_val is None:
            continue  # truly empty row

        try:
            txn_date = _xlsx_cell_to_date(date_val)
        except ValueError as exc:
            validation_errors.append(f"Row {row_num}: invalid date — {exc}")
            continue

        try:
            valor_float = _xlsx_cell_to_float(valor_val)
        except ValueError as exc:
            validation_errors.append(f"Row {row_num}: invalid valor — {exc}")
            continue

        amount_cents = round(abs(valor_float) * 100)
        if amount_cents == 0:
            validation_errors.append(f"Row {row_num}: zero-value transaction skipped")
            continue

        # Map Tipo → internal type
        transfer_str = str(transferencia_val).strip() if transferencia_val else ""

        if tipo == "Despesa":
            tx_type = "expense"
            transfer_to = None
        elif tipo == "Receita":
            tx_type = "income"
            transfer_to = None
        elif tipo == "Transferência":
            # Only import the transfer from the SOURCE account's sheet.
            # When current bank is the DESTINATION, the router will update
            # this account's balance automatically when the source-side
            # transfer is imported from the other bank's sheet.
            if not _is_transfer_source(transfer_str, bank_key):
                # Destination side — track balance but skip as a transaction
                calculated_balance += valor_float
                continue
            tx_type = "transfer"
            transfer_to = _resolve_transfer_dest(
                transfer_str, itau_account_id, inter_account_id
            )
        else:
            # Unknown tipo — skip with warning
            validation_errors.append(
                f"Row {row_num}: unknown Tipo '{tipo}', skipped"
            )
            continue

        # Running balance using raw valor sign (sheet is reconciled)
        calculated_balance += valor_float

        # Build notes: store original categoria and ordem for auditability
        notes_parts: list[str] = []
        if categoria_val:
            notes_parts.append(f"Categoria: {categoria_val}")
        if ordem_val is not None:
            notes_parts.append(f"Ordem: {ordem_val}")
        if transferencia_val:
            notes_parts.append(f"Transferência: {transferencia_val}")

        description = str(descricao_val).strip()[:200] if descricao_val else ""

        txn: dict[str, Any] = {
            "account_id": account_id,
            "category_id": None,  # caller can resolve by name from 'notes'
            "amount_cents": amount_cents,
            "type": tx_type,
            "description": description,
            "date": txn_date,
            "notes": " | ".join(notes_parts) or None,
            "transfer_to_account_id": transfer_to,
            "is_reconciled": True,
            "source": "import",
        }
        transactions.append(txn)

    calculated_final_cents = round(calculated_balance * 100)

    # Validate: calculated balance must match expected final balance
    tolerance = 1  # allow 1-cent rounding difference
    if abs(calculated_final_cents - expected_final_cents) > tolerance:
        validation_errors.append(
            f"Balance mismatch: calculated R$ {calculated_final_cents / 100:.2f} "
            f"vs expected R$ {expected_final_cents / 100:.2f} "
            f"(diff = {calculated_final_cents - expected_final_cents} cents)"
        )

    return {
        "rows": transactions,
        "initial_balance_cents": initial_balance_cents,
        "expected_final_cents": expected_final_cents,
        "calculated_final_cents": calculated_final_cents,
        "validation_errors": validation_errors,
    }


def parse_xlsx_reconciled(
    content: bytes,
    itau_account_id: str,
    inter_account_id: str,
) -> dict[str, Any]:
    """
    Parse the reconciled XLSX spreadsheet (lancamento_por_lancamento_conferido.xlsx).

    Processes sheets 'Itaú' and 'Inter' separately.
    Skips 'Conferência final' and 'Todos os lançamentos' (audit-only).

    Args:
        content: Raw xlsx bytes.
        itau_account_id: UUID of the Itaú account in the database.
        inter_account_id: UUID of the Inter account in the database.

    Returns:
        {
            "itau": { "rows": [...], "initial_balance_cents": int,
                      "expected_final_cents": int, "calculated_final_cents": int,
                      "validation_errors": [...] },
            "inter": { ... },
        }
    """
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    result: dict[str, Any] = {}

    for sheet_title, bank_key in _XLSX_BANK_SHEETS.items():
        # Tolerate minor name variations (e.g. trailing spaces, accent encoding)
        ws = None
        for name in wb.sheetnames:
            if name.strip().lower() == sheet_title.lower():
                ws = wb[name]
                break

        if ws is None:
            raise ValueError(
                f"Sheet '{sheet_title}' not found in xlsx. "
                f"Available sheets: {wb.sheetnames}"
            )

        account_id = itau_account_id if bank_key == "itau" else inter_account_id

        result[bank_key] = _parse_xlsx_sheet(
            ws, account_id, bank_key, itau_account_id, inter_account_id
        )

    return result


def parse_file(
    content: bytes, filename: str, account_id: str
) -> list[dict[str, Any]]:
    """
    Auto-detect format from filename extension and content, parse, inject account_id.

    Args:
        content: Raw file bytes.
        filename: Original filename (used for format detection).
        account_id: UUID of the account to assign transactions to.

    Returns:
        List of transaction dicts ready for bulk insert.
    """
    lower = filename.lower()
    rows: list[dict[str, Any]]

    if lower.endswith(".ofx") or lower.endswith(".qfx"):
        rows = parse_ofx(content)
    elif lower.endswith(".csv"):
        # Detect bank from first 512 bytes
        header_sample = content[:512].decode("utf-8-sig", errors="ignore")
        bank = _detect_bank(header_sample)
        if bank == "nubank":
            rows = parse_nubank_csv(content)
        elif bank == "inter":
            rows = parse_inter_csv(content)
        elif bank in ("bb", "banco do brasil"):
            rows = parse_bb_csv(content)
        else:
            # Default: try Nubank layout first, fall back to Inter
            try:
                rows = parse_nubank_csv(content)
            except Exception:
                rows = parse_inter_csv(content)
    else:
        raise ValueError(f"Unsupported file format: {filename!r}")

    for row in rows:
        row["account_id"] = account_id

    return rows
